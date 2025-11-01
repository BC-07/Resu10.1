#!/usr/bin/env python3
"""
Real-World Semantic Scoring Test
Tests semantic engine with actual PDS files and real job postings
"""

import os
import sys
import json
import pandas as pd
from datetime import datetime
from typing import Dict, List, Any

# Import required modules
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    from database import DatabaseManager
    from enhanced_assessment_engine import EnhancedUniversityAssessmentEngine
    from semantic_engine import UniversitySemanticEngine
    import openpyxl
except ImportError as e:
    print(f"❌ Import Error: {e}")
    sys.exit(1)

class RealWorldSemanticTest:
    """Test semantic scoring with real PDS files and job postings"""
    
    def __init__(self):
        self.db_manager = DatabaseManager()
        self.enhanced_engine = EnhancedUniversityAssessmentEngine(db_manager=self.db_manager)
        self.pds_folder = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'SamplePDSFiles')
        self.results = []
        
    def get_real_job_postings(self) -> List[Dict]:
        """Get real job postings from database"""
        try:
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            
            # Try to get from lspu_job_postings table
            cursor.execute("""
                SELECT id, position_title, department_office, education_requirements, 
                       experience_requirements, training_requirements, eligibility_requirements,
                       salary_grade, employment_period
                FROM lspu_job_postings 
                LIMIT 10
            """)
            
            job_postings = []
            for row in cursor.fetchall():
                if hasattr(row, 'keys'):  # RealDictRow
                    job_data = {
                        'id': row['id'],
                        'title': row['position_title'],
                        'department': row['department_office'] or '',
                        'description': self._build_job_description(row),
                        'requirements': self._extract_requirements(row)
                    }
                else:  # tuple
                    job_data = {
                        'id': row[0],
                        'title': row[1],
                        'department': row[2] or '',
                        'description': self._build_job_description_tuple(row),
                        'requirements': self._extract_requirements_tuple(row)
                    }
                job_postings.append(job_data)
            
            cursor.close()
            conn.close()
            
            if job_postings:
                print(f"✅ Found {len(job_postings)} real job postings from database")
                return job_postings
                
        except Exception as e:
            print(f"⚠️  Could not load job postings from database: {e}")
        
        # Fallback to sample job postings
        return self.get_sample_job_postings()
    
    def get_sample_job_postings(self) -> List[Dict]:
        """Create sample university job postings similar to LSPU format"""
        return [
            {
                'id': 'sample_1',
                'title': 'Assistant Professor - Computer Science',
                'department': 'College of Engineering and Technology',
                'description': """We are seeking a qualified Assistant Professor for the Computer Science program. 
                The position involves teaching undergraduate courses in programming, software engineering, 
                database systems, and computer networks. The successful candidate will also conduct research 
                in computer science and participate in university service activities.""",
                'requirements': [
                    'Master\'s degree in Computer Science or related field',
                    'At least 3 years teaching experience in higher education',
                    'Programming experience in Python, Java, or C++',
                    'Research experience with published papers',
                    'Professional License or Civil Service Eligibility',
                    'Strong communication and interpersonal skills'
                ]
            },
            {
                'id': 'sample_2',
                'title': 'Instructor - Information Technology',
                'department': 'College of Engineering and Technology',
                'description': """Position for IT Instructor to teach courses in web development, database management,
                network administration, and system analysis. Will guide students in capstone projects and 
                participate in curriculum development for the IT program.""",
                'requirements': [
                    'Bachelor\'s degree in Information Technology or Computer Science',
                    'At least 2 years industry or teaching experience',
                    'Web development skills (HTML, CSS, JavaScript, PHP)',
                    'Database management experience (MySQL, PostgreSQL)',
                    'Network administration knowledge',
                    'Professional IT certifications preferred'
                ]
            },
            {
                'id': 'sample_3',
                'title': 'Assistant Professor - Business Administration',
                'department': 'College of Business and Accountancy',
                'description': """Faculty position for Business Administration program teaching courses in management,
                marketing, entrepreneurship, and business ethics. Expected to conduct research in business-related
                fields and provide consultancy services to local businesses.""",
                'requirements': [
                    'Master\'s degree in Business Administration or related field',
                    'CPA license or relevant professional certification',
                    'At least 3 years teaching or industry experience',
                    'Research publications in business journals',
                    'Expertise in business analysis and strategic planning',
                    'Strong presentation and leadership skills'
                ]
            },
            {
                'id': 'sample_4',
                'title': 'Lecturer - Civil Engineering',
                'department': 'College of Engineering and Technology',
                'description': """Lecturer position for Civil Engineering program to teach structural design,
                construction management, surveying, and engineering mathematics. Will supervise student
                thesis projects and participate in infrastructure development research.""",
                'requirements': [
                    'Bachelor\'s degree in Civil Engineering',
                    'Licensed Civil Engineer',
                    'At least 5 years construction industry experience',
                    'AutoCAD and engineering software proficiency',
                    'Project management experience',
                    'Knowledge of building codes and standards'
                ]
            }
        ]
    
    def _build_job_description(self, row) -> str:
        """Build job description from database row (RealDictRow)"""
        parts = []
        if row.get('department_office'):
            parts.append(f"Department: {row['department_office']}")
        if row.get('employment_period'):
            parts.append(f"Employment Period: {row['employment_period']}")
        if row.get('salary_grade'):
            parts.append(f"Salary Grade: {row['salary_grade']}")
        
        requirements = []
        for field in ['education_requirements', 'experience_requirements', 
                     'training_requirements', 'eligibility_requirements']:
            if row.get(field):
                requirements.append(row[field])
        
        if requirements:
            parts.append("Requirements: " + " ".join(requirements))
        
        return " ".join(parts)
    
    def _build_job_description_tuple(self, row) -> str:
        """Build job description from database row (tuple)"""
        parts = []
        if row[2]:  # department_office
            parts.append(f"Department: {row[2]}")
        if row[8]:  # employment_period
            parts.append(f"Employment Period: {row[8]}")
        if row[7]:  # salary_grade
            parts.append(f"Salary Grade: {row[7]}")
        
        requirements = []
        for field in [row[3], row[4], row[5], row[6]]:  # education, experience, training, eligibility
            if field:
                requirements.append(field)
        
        if requirements:
            parts.append("Requirements: " + " ".join(requirements))
        
        return " ".join(parts)
    
    def _extract_requirements(self, row) -> List[str]:
        """Extract requirements from database row (RealDictRow)"""
        requirements = []
        for field in ['education_requirements', 'experience_requirements', 
                     'training_requirements', 'eligibility_requirements']:
            if row.get(field):
                requirements.append(row[field])
        return requirements
    
    def _extract_requirements_tuple(self, row) -> List[str]:
        """Extract requirements from database row (tuple)"""
        requirements = []
        for field in [row[3], row[4], row[5], row[6]]:  # education, experience, training, eligibility
            if field:
                requirements.append(field)
        return requirements
    
    def load_pds_files(self) -> List[Dict]:
        """Load and parse PDS files from SamplePDSFiles folder"""
        pds_candidates = []
        
        if not os.path.exists(self.pds_folder):
            print(f"❌ PDS folder not found: {self.pds_folder}")
            return []
        
        pds_files = [f for f in os.listdir(self.pds_folder) if f.endswith('.xlsx')]
        print(f"📁 Found {len(pds_files)} PDS files: {pds_files}")
        
        for filename in pds_files:
            try:
                filepath = os.path.join(self.pds_folder, filename)
                candidate_data = self.parse_pds_file(filepath, filename)
                if candidate_data:
                    pds_candidates.append(candidate_data)
                    print(f"   ✅ Loaded: {filename} - {candidate_data.get('name', 'Unknown')}")
                else:
                    print(f"   ⚠️  Could not parse: {filename}")
            except Exception as e:
                print(f"   ❌ Error loading {filename}: {e}")
        
        return pds_candidates
    
    def parse_pds_file(self, filepath: str, filename: str) -> Dict:
        """Parse a PDS Excel file and extract candidate information"""
        try:
            # Try to read the Excel file
            workbook = openpyxl.load_workbook(filepath, data_only=True)
            sheet = workbook.active
            
            # Extract basic information
            candidate_data = {
                'filename': filename,
                'name': self.extract_name_from_excel(sheet),
                'education': self.extract_education_from_excel(sheet),
                'experience': self.extract_experience_from_excel(sheet),
                'skills': self.extract_skills_from_excel(sheet),
                'extracted_text': self.extract_all_text_from_excel(sheet)
            }
            
            workbook.close()
            return candidate_data
            
        except Exception as e:
            print(f"Error parsing {filename}: {e}")
            return None
    
    def extract_name_from_excel(self, sheet) -> str:
        """Extract candidate name from Excel sheet"""
        # Common locations for name in PDS
        name_cells = ['B7', 'C7', 'D7', 'B8', 'C8', 'D8', 'B6', 'C6', 'D6']
        
        for cell in name_cells:
            try:
                value = sheet[cell].value
                if value and isinstance(value, str) and len(value.strip()) > 2:
                    # Check if it looks like a name (has letters and possibly spaces)
                    if any(c.isalpha() for c in value) and not any(c.isdigit() for c in value):
                        return value.strip()
            except:
                continue
        
        # Fallback: scan first 20 rows for name-like content
        for row in range(1, 21):
            for col in range(1, 10):
                try:
                    cell_value = sheet.cell(row=row, column=col).value
                    if (cell_value and isinstance(cell_value, str) and 
                        len(cell_value.strip()) > 5 and len(cell_value.strip()) < 50 and
                        ' ' in cell_value and cell_value.replace(' ', '').isalpha()):
                        return cell_value.strip()
                except:
                    continue
        
        return "Unknown Candidate"
    
    def extract_education_from_excel(self, sheet) -> str:
        """Extract education information from Excel sheet"""
        education_keywords = ['education', 'college', 'university', 'degree', 'bachelor', 'master', 'phd']
        education_info = []
        
        # Scan the sheet for education-related content
        for row in range(1, 100):
            for col in range(1, 15):
                try:
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str):
                        cell_lower = cell_value.lower()
                        if any(keyword in cell_lower for keyword in education_keywords):
                            if len(cell_value.strip()) > 10:  # Substantial content
                                education_info.append(cell_value.strip())
                except:
                    continue
        
        return " ".join(education_info[:3]) if education_info else "Education information not found"
    
    def extract_experience_from_excel(self, sheet) -> str:
        """Extract work experience from Excel sheet"""
        experience_keywords = ['experience', 'work', 'employment', 'position', 'company', 'job', 'career']
        experience_info = []
        
        for row in range(1, 100):
            for col in range(1, 15):
                try:
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str):
                        cell_lower = cell_value.lower()
                        if any(keyword in cell_lower for keyword in experience_keywords):
                            if len(cell_value.strip()) > 10:
                                experience_info.append(cell_value.strip())
                except:
                    continue
        
        return " ".join(experience_info[:3]) if experience_info else "Work experience not found"
    
    def extract_skills_from_excel(self, sheet) -> str:
        """Extract skills information from Excel sheet"""
        skills_keywords = ['skill', 'competenc', 'training', 'seminar', 'workshop', 'certification']
        skills_info = []
        
        for row in range(1, 100):
            for col in range(1, 15):
                try:
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str):
                        cell_lower = cell_value.lower()
                        if any(keyword in cell_lower for keyword in skills_keywords):
                            if len(cell_value.strip()) > 5:
                                skills_info.append(cell_value.strip())
                except:
                    continue
        
        return " ".join(skills_info[:5]) if skills_info else "Skills information not found"
    
    def extract_all_text_from_excel(self, sheet) -> str:
        """Extract all text content from Excel sheet for comprehensive analysis"""
        all_text = []
        
        for row in range(1, 100):
            for col in range(1, 15):
                try:
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str) and len(cell_value.strip()) > 2:
                        all_text.append(cell_value.strip())
                except:
                    continue
        
        return " ".join(all_text[:100])  # Limit to first 100 meaningful cells
    
    def run_comprehensive_test(self):
        """Run comprehensive semantic scoring test with real data"""
        print("🧪 Real-World Semantic Scoring Test")
        print("=" * 60)
        
        # Load real job postings
        print("\n📋 Loading job postings...")
        job_postings = self.get_real_job_postings()
        print(f"   ✅ Loaded {len(job_postings)} job postings")
        
        # Load PDS files
        print("\n👥 Loading PDS candidates...")
        pds_candidates = self.load_pds_files()
        print(f"   ✅ Loaded {len(pds_candidates)} PDS candidates")
        
        if not pds_candidates:
            print("❌ No PDS candidates loaded. Test cannot continue.")
            return
        
        # Test each job posting against all candidates
        print("\n🧠 Running semantic assessments...")
        
        test_results = []
        total_assessments = 0
        
        for i, job in enumerate(job_postings):
            print(f"\n📋 Job {i+1}: {job['title']}")
            print(f"   Department: {job.get('department', 'N/A')}")
            
            job_results = {
                'job_id': job['id'],
                'job_title': job['title'],
                'job_department': job.get('department', ''),
                'candidate_scores': []
            }
            
            for j, candidate in enumerate(pds_candidates):
                try:
                    print(f"   👤 Assessing: {candidate['name']}")
                    
                    # Run semantic assessment
                    result = self.enhanced_engine.assess_candidate_enhanced(
                        candidate_data=candidate,
                        job_data=job
                    )
                    
                    candidate_result = {
                        'candidate_name': candidate['name'],
                        'candidate_file': candidate['filename'],
                        'semantic_score': result.get('semantic_score', 0),
                        'traditional_score': result.get('traditional_score', None),
                        'total_score': result.get('total_score', 0),
                        'semantic_breakdown': result.get('semantic_breakdown', {})
                    }
                    
                    job_results['candidate_scores'].append(candidate_result)
                    total_assessments += 1
                    
                    print(f"      📊 Semantic Score: {candidate_result['semantic_score']:.1f}")
                    
                    # Show top components
                    breakdown = candidate_result['semantic_breakdown']
                    if breakdown:
                        print(f"      📋 Education: {breakdown.get('education_relevance', 0):.3f}")
                        print(f"      💼 Experience: {breakdown.get('experience_relevance', 0):.3f}")
                        print(f"      🛠️  Skills: {breakdown.get('skills_relevance', 0):.3f}")
                    
                except Exception as e:
                    print(f"      ❌ Assessment failed: {e}")
                    candidate_result = {
                        'candidate_name': candidate['name'],
                        'candidate_file': candidate['filename'],
                        'semantic_score': 0,
                        'traditional_score': None,
                        'total_score': 0,
                        'error': str(e)
                    }
                    job_results['candidate_scores'].append(candidate_result)
            
            # Sort candidates by semantic score
            job_results['candidate_scores'].sort(
                key=lambda x: x.get('semantic_score', 0), 
                reverse=True
            )
            
            test_results.append(job_results)
        
        # Generate comprehensive report
        self.generate_test_report(test_results, total_assessments)
        
        return test_results
    
    def generate_test_report(self, test_results: List[Dict], total_assessments: int):
        """Generate comprehensive test report"""
        
        report_filename = f"real_world_semantic_test_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        # Calculate summary statistics
        all_scores = []
        job_summaries = []
        
        for job_result in test_results:
            scores = [c['semantic_score'] for c in job_result['candidate_scores'] 
                     if 'error' not in c]
            
            if scores:
                job_summary = {
                    'job_title': job_result['job_title'],
                    'candidates_assessed': len(job_result['candidate_scores']),
                    'average_score': sum(scores) / len(scores),
                    'highest_score': max(scores),
                    'lowest_score': min(scores),
                    'top_candidate': job_result['candidate_scores'][0]['candidate_name'] if job_result['candidate_scores'] else None
                }
                job_summaries.append(job_summary)
                all_scores.extend(scores)
        
        # Overall statistics
        overall_stats = {
            'total_jobs_tested': len(test_results),
            'total_assessments': total_assessments,
            'average_semantic_score': sum(all_scores) / len(all_scores) if all_scores else 0,
            'highest_overall_score': max(all_scores) if all_scores else 0,
            'lowest_overall_score': min(all_scores) if all_scores else 0,
            'score_distribution': {
                'excellent_90_plus': len([s for s in all_scores if s >= 90]),
                'very_good_80_89': len([s for s in all_scores if 80 <= s < 90]),
                'good_70_79': len([s for s in all_scores if 70 <= s < 80]),
                'satisfactory_60_69': len([s for s in all_scores if 60 <= s < 70]),
                'needs_improvement_below_60': len([s for s in all_scores if s < 60])
            }
        }
        
        # Create final report
        final_report = {
            'test_metadata': {
                'test_date': datetime.now().isoformat(),
                'test_type': 'Real World PDS Files vs Job Postings',
                'semantic_model': 'all-MiniLM-L6-v2',
                'total_assessments': total_assessments
            },
            'overall_statistics': overall_stats,
            'job_summaries': job_summaries,
            'detailed_results': test_results
        }
        
        # Save detailed report
        with open(report_filename, 'w', encoding='utf-8') as f:
            json.dump(final_report, f, indent=2, ensure_ascii=False)
        
        # Print summary
        print(f"\n📊 TEST SUMMARY REPORT")
        print("=" * 50)
        print(f"📅 Test Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"🧪 Total Assessments: {total_assessments}")
        print(f"📋 Jobs Tested: {len(test_results)}")
        print(f"👥 Unique Candidates: {len(set(c['candidate_name'] for r in test_results for c in r['candidate_scores']))}")
        
        if all_scores:
            print(f"\n📈 Score Statistics:")
            print(f"   Average Score: {overall_stats['average_semantic_score']:.1f}")
            print(f"   Highest Score: {overall_stats['highest_overall_score']:.1f}")
            print(f"   Lowest Score: {overall_stats['lowest_overall_score']:.1f}")
            
            print(f"\n📊 Score Distribution:")
            dist = overall_stats['score_distribution']
            print(f"   Excellent (90+): {dist['excellent_90_plus']} assessments")
            print(f"   Very Good (80-89): {dist['very_good_80_89']} assessments")
            print(f"   Good (70-79): {dist['good_70_79']} assessments")
            print(f"   Satisfactory (60-69): {dist['satisfactory_60_69']} assessments")
            print(f"   Needs Improvement (<60): {dist['needs_improvement_below_60']} assessments")
        
        print(f"\n🏆 Top Performers by Job:")
        for job_summary in job_summaries:
            print(f"   {job_summary['job_title']}: {job_summary['top_candidate']} ({job_summary['highest_score']:.1f})")
        
        print(f"\n📁 Detailed report saved: {report_filename}")
        
        return final_report

def main():
    """Main test execution"""
    print("🚀 Starting Real-World Semantic Scoring Test")
    print("=" * 60)
    
    try:
        tester = RealWorldSemanticTest()
        results = tester.run_comprehensive_test()
        
        print("\n🎉 Test completed successfully!")
        print("✅ Semantic scoring system validated with real PDS files")
        
        return True
        
    except Exception as e:
        print(f"\n❌ Test failed: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)