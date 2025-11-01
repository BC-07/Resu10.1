#!/usr/bin/env python3
"""
Improved Real-World Semantic Test with Better PDS Parsing
Enhanced version with detailed diagnostics and improved PDS file parsing
"""

import os
import sys
import json
import pandas as pd
from datetime import datetime
from typing import Dict, List, Any

# Import required modules
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    from database import DatabaseManager
    from enhanced_assessment_engine import EnhancedUniversityAssessmentEngine
    from semantic_engine import UniversitySemanticEngine
    import openpyxl
except ImportError as e:
    print(f"❌ Import Error: {e}")
    sys.exit(1)

class ImprovedPDSParser:
    """Improved PDS file parser with better content extraction"""
    
    def __init__(self):
        self.name_indicators = [
            'surname', 'first name', 'middle name', 'name', 'apellido', 'nombre'
        ]
        self.education_indicators = [
            'elementary', 'high school', 'college', 'university', 'bachelor', 'master',
            'phd', 'doctorate', 'degree', 'diploma', 'certificate', 'school attended'
        ]
        self.experience_indicators = [
            'position title', 'company', 'office', 'employment', 'work experience',
            'position', 'job', 'employer', 'agency', 'department'
        ]
        self.skills_indicators = [
            'training', 'seminar', 'workshop', 'skills', 'competencies', 'certification',
            'course', 'program attended', 'special skills'
        ]
    
    def parse_pds_file(self, filepath: str, filename: str) -> Dict:
        """Enhanced PDS parsing with better content extraction"""
        try:
            workbook = openpyxl.load_workbook(filepath, data_only=True)
            sheet = workbook.active
            
            # Get all cell data first
            all_cells = self.extract_all_cells(sheet)
            
            # Parse specific sections
            candidate_data = {
                'filename': filename,
                'name': self.extract_name_improved(all_cells),
                'education': self.extract_education_improved(all_cells),
                'experience': self.extract_experience_improved(all_cells),
                'skills': self.extract_skills_improved(all_cells),
                'extracted_text': self.extract_comprehensive_text(all_cells),
                'raw_content_preview': self.get_content_preview(all_cells)
            }
            
            workbook.close()
            return candidate_data
            
        except Exception as e:
            print(f"Error parsing {filename}: {e}")
            return None
    
    def extract_all_cells(self, sheet) -> List[Dict]:
        """Extract all cell data with position information"""
        cells = []
        
        for row in range(1, min(200, sheet.max_row + 1)):  # Limit to 200 rows
            for col in range(1, min(20, sheet.max_column + 1)):  # Limit to 20 columns
                try:
                    cell = sheet.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str) and len(cell.value.strip()) > 1:
                        cells.append({
                            'row': row,
                            'col': col,
                            'value': cell.value.strip(),
                            'value_lower': cell.value.strip().lower()
                        })
                except:
                    continue
        
        return cells
    
    def extract_name_improved(self, cells: List[Dict]) -> str:
        """Improved name extraction with pattern matching"""
        # Look for name patterns
        potential_names = []
        
        for cell in cells:
            value = cell['value']
            
            # Skip obvious non-names
            if (any(word in cell['value_lower'] for word in ['yes', 'no', 'male', 'female', 'single', 'married']) or
                len(value) < 3 or len(value) > 60 or
                any(char.isdigit() for char in value) or
                value.lower() in ['n/a', 'na', 'none', 'not applicable']):
                continue
            
            # Look for name patterns
            if ((' ' in value and value.replace(' ', '').replace('.', '').isalpha()) and
                len(value.split()) >= 2 and len(value.split()) <= 4):
                
                # Check if it's likely a name (title case, multiple words)
                words = value.split()
                if all(len(word) >= 2 for word in words) and any(word[0].isupper() for word in words):
                    potential_names.append({
                        'name': value,
                        'position': (cell['row'], cell['col']),
                        'confidence': self.calculate_name_confidence(value, cell)
                    })
        
        # Sort by confidence and return the best match
        if potential_names:
            best_name = max(potential_names, key=lambda x: x['confidence'])
            if best_name['confidence'] > 0.5:
                return best_name['name']
        
        # Fallback to filename-based name
        return f"Candidate from {cells[0]['value'] if cells else 'Unknown File'}"
    
    def calculate_name_confidence(self, name: str, cell: Dict) -> float:
        """Calculate confidence score for a potential name"""
        confidence = 0.0
        
        # Position scoring (names usually in top rows)
        if cell['row'] <= 20:
            confidence += 0.3
        
        # Word count scoring (2-4 words typical for names)
        word_count = len(name.split())
        if 2 <= word_count <= 4:
            confidence += 0.4
        
        # Case scoring (proper names usually title case)
        if name.istitle():
            confidence += 0.3
        
        # Length scoring
        if 8 <= len(name) <= 40:
            confidence += 0.2
        
        return min(confidence, 1.0)
    
    def extract_education_improved(self, cells: List[Dict]) -> str:
        """Improved education extraction"""
        education_content = []
        
        for cell in cells:
            if any(indicator in cell['value_lower'] for indicator in self.education_indicators):
                if len(cell['value']) > 10 and len(cell['value']) < 200:
                    education_content.append(cell['value'])
        
        # Also look for degree-like patterns
        for cell in cells:
            value = cell['value']
            if (('bachelor' in cell['value_lower'] or 'master' in cell['value_lower'] or
                 'bs' in cell['value_lower'] or 'ms' in cell['value_lower'] or
                 'phd' in cell['value_lower'] or 'doctorate' in cell['value_lower']) and
                len(value) > 5):
                education_content.append(value)
        
        return ' | '.join(education_content[:5]) if education_content else "Education not clearly specified"
    
    def extract_experience_improved(self, cells: List[Dict]) -> str:
        """Improved work experience extraction"""
        experience_content = []
        
        for cell in cells:
            if any(indicator in cell['value_lower'] for indicator in self.experience_indicators):
                if len(cell['value']) > 8 and len(cell['value']) < 200:
                    experience_content.append(cell['value'])
        
        # Look for job title patterns
        job_indicators = ['engineer', 'manager', 'supervisor', 'assistant', 'coordinator', 
                         'officer', 'specialist', 'analyst', 'director', 'administrator']
        
        for cell in cells:
            if any(job in cell['value_lower'] for job in job_indicators):
                if len(cell['value']) > 5 and len(cell['value']) < 100:
                    experience_content.append(cell['value'])
        
        return ' | '.join(experience_content[:5]) if experience_content else "Work experience not clearly specified"
    
    def extract_skills_improved(self, cells: List[Dict]) -> str:
        """Improved skills extraction"""
        skills_content = []
        
        for cell in cells:
            if any(indicator in cell['value_lower'] for indicator in self.skills_indicators):
                if len(cell['value']) > 5 and len(cell['value']) < 200:
                    skills_content.append(cell['value'])
        
        # Look for technical skills
        tech_indicators = ['computer', 'software', 'programming', 'microsoft', 'excel', 
                          'word', 'powerpoint', 'database', 'web', 'internet']
        
        for cell in cells:
            if any(tech in cell['value_lower'] for tech in tech_indicators):
                if len(cell['value']) > 5 and len(cell['value']) < 150:
                    skills_content.append(cell['value'])
        
        return ' | '.join(skills_content[:8]) if skills_content else "Skills not clearly specified"
    
    def extract_comprehensive_text(self, cells: List[Dict]) -> str:
        """Extract comprehensive text for semantic analysis"""
        # Filter out obvious form fields and short entries
        meaningful_content = []
        
        for cell in cells:
            value = cell['value']
            
            # Skip obvious form fields
            if (len(value) < 3 or 
                value.lower() in ['yes', 'no', 'n/a', 'na', 'none', 'male', 'female'] or
                value.lower().startswith('date') or
                all(char in '()[]/-:.,; ' for char in value.replace(value.strip(), '')) or
                len(value) > 500):  # Skip very long entries
                continue
            
            # Include substantial content
            if len(value) >= 3:
                meaningful_content.append(value)
        
        # Limit to reasonable amount of text for processing
        return ' '.join(meaningful_content[:150])
    
    def get_content_preview(self, cells: List[Dict]) -> str:
        """Get a preview of the extracted content for debugging"""
        preview_items = []
        for cell in cells[:20]:  # First 20 meaningful cells
            preview_items.append(f"[{cell['row']},{cell['col']}]: {cell['value'][:50]}")
        return ' | '.join(preview_items)

class DiagnosticSemanticTest:
    """Enhanced semantic test with detailed diagnostics"""
    
    def __init__(self):
        self.db_manager = DatabaseManager()
        self.enhanced_engine = EnhancedUniversityAssessmentEngine(db_manager=self.db_manager)
        self.semantic_engine = UniversitySemanticEngine()
        self.pds_parser = ImprovedPDSParser()
        self.pds_folder = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'SamplePDSFiles')
    
    def load_improved_pds_files(self) -> List[Dict]:
        """Load PDS files with improved parsing"""
        pds_candidates = []
        
        if not os.path.exists(self.pds_folder):
            print(f"❌ PDS folder not found: {self.pds_folder}")
            return []
        
        pds_files = [f for f in os.listdir(self.pds_folder) if f.endswith('.xlsx')]
        print(f"📁 Found {len(pds_files)} PDS files: {pds_files}")
        
        for filename in pds_files:
            try:
                filepath = os.path.join(self.pds_folder, filename)
                candidate_data = self.pds_parser.parse_pds_file(filepath, filename)
                
                if candidate_data:
                    pds_candidates.append(candidate_data)
                    print(f"   ✅ Loaded: {filename}")
                    print(f"      👤 Name: {candidate_data['name']}")
                    print(f"      🎓 Education: {candidate_data['education'][:80]}...")
                    print(f"      💼 Experience: {candidate_data['experience'][:80]}...")
                    print(f"      🛠️  Skills: {candidate_data['skills'][:80]}...")
                else:
                    print(f"   ⚠️  Could not parse: {filename}")
                    
            except Exception as e:
                print(f"   ❌ Error loading {filename}: {e}")
        
        return pds_candidates
    
    def get_realistic_job_postings(self) -> List[Dict]:
        """Get realistic university job postings for testing"""
        return [
            {
                'id': 'prof_cs',
                'title': 'Assistant Professor - Computer Science',
                'department': 'College of Engineering and Technology',
                'description': """The Computer Science Department seeks an Assistant Professor to teach undergraduate 
                courses in programming, data structures, algorithms, software engineering, and database systems. 
                Responsibilities include conducting research in computer science, mentoring students, and participating 
                in department service activities. The ideal candidate should have strong programming skills and 
                experience in academic research.""",
                'requirements': [
                    'Master\'s or PhD in Computer Science or closely related field',
                    'At least 3 years of teaching experience in higher education',
                    'Programming experience in Java, Python, C++, or similar languages',
                    'Research publications in computer science journals or conferences',
                    'Experience with database systems and web development',
                    'Professional license or civil service eligibility',
                    'Strong communication and interpersonal skills'
                ]
            },
            {
                'id': 'lecturer_it',
                'title': 'Lecturer - Information Technology',
                'department': 'College of Engineering and Technology', 
                'description': """Position for IT Lecturer to teach undergraduate courses in web development, 
                network administration, systems analysis, and IT project management. The role involves curriculum 
                development, student mentoring, and collaboration with industry partners. Candidates should have 
                practical IT experience and academic qualifications.""",
                'requirements': [
                    'Bachelor\'s degree in Information Technology, Computer Science, or related field',
                    'At least 5 years of industry experience in IT or 3 years teaching experience',
                    'Web development skills (HTML, CSS, JavaScript, PHP, or similar)',
                    'Network administration and cybersecurity knowledge',
                    'Database management experience (MySQL, PostgreSQL, Oracle)',
                    'IT certifications (Cisco, Microsoft, CompTIA) preferred',
                    'Project management experience'
                ]
            },
            {
                'id': 'prof_business',
                'title': 'Associate Professor - Business Administration',
                'department': 'College of Business and Accountancy',
                'description': """The Business Administration Department seeks an Associate Professor to teach 
                graduate and undergraduate courses in management, marketing, entrepreneurship, and strategic planning. 
                The position requires conducting research in business-related fields, supervising thesis students, 
                and providing consultancy services to local businesses and organizations.""",
                'requirements': [
                    'PhD in Business Administration, Management, or related field',
                    'At least 7 years of combined teaching and industry experience',
                    'CPA license or MBA preferred',
                    'Research publications in business or management journals',
                    'Experience in business consulting or management',
                    'Leadership and team management skills',
                    'Expertise in strategic planning and business analysis'
                ]
            }
        ]
    
    def run_diagnostic_test(self):
        """Run comprehensive diagnostic test"""
        print("🔍 Diagnostic Real-World Semantic Test")
        print("=" * 60)
        
        # Load data
        print("\n📋 Loading test data...")
        job_postings = self.get_realistic_job_postings()
        pds_candidates = self.load_improved_pds_files()
        
        if not pds_candidates:
            print("❌ No candidates loaded. Cannot proceed with test.")
            return
        
        print(f"\n✅ Test Setup Complete:")
        print(f"   📋 Job Postings: {len(job_postings)}")
        print(f"   👥 PDS Candidates: {len(pds_candidates)}")
        
        # Run detailed assessments
        print(f"\n🧠 Running detailed semantic assessments...")
        
        detailed_results = []
        
        for i, job in enumerate(job_postings):
            print(f"\n📋 Job {i+1}: {job['title']}")
            print(f"   🏢 Department: {job['department']}")
            
            job_results = {
                'job_info': job,
                'candidate_assessments': []
            }
            
            for j, candidate in enumerate(pds_candidates):
                print(f"\n   👤 Candidate {j+1}: {candidate['name']}")
                
                try:
                    # Get detailed semantic breakdown
                    semantic_details = self.semantic_engine.calculate_detailed_semantic_score(
                        job_data=job,
                        candidate_data=candidate
                    )
                    
                    # Run full enhanced assessment
                    assessment_result = self.enhanced_engine.assess_candidate_enhanced(
                        candidate_data=candidate,
                        job_data=job
                    )
                    
                    candidate_assessment = {
                        'candidate_info': {
                            'name': candidate['name'],
                            'filename': candidate['filename'],
                            'education_preview': candidate['education'][:100],
                            'experience_preview': candidate['experience'][:100],
                            'skills_preview': candidate['skills'][:100]
                        },
                        'semantic_details': semantic_details,
                        'assessment_result': assessment_result,
                        'diagnostic_info': self.generate_diagnostic_info(semantic_details, candidate, job)
                    }
                    
                    job_results['candidate_assessments'].append(candidate_assessment)
                    
                    # Print detailed results
                    print(f"      📊 Semantic Score: {assessment_result.get('semantic_score', 0):.1f}")
                    print(f"      📋 Education Relevance: {semantic_details.get('education_relevance', 0):.3f}")
                    print(f"      💼 Experience Relevance: {semantic_details.get('experience_relevance', 0):.3f}")
                    print(f"      🛠️  Skills Relevance: {semantic_details.get('skills_relevance', 0):.3f}")
                    print(f"      🌟 Overall Similarity: {semantic_details.get('overall_similarity', 0):.3f}")
                    
                    # Show diagnostic insights
                    diagnostic = candidate_assessment['diagnostic_info']
                    print(f"      💡 Match Quality: {diagnostic['match_quality']}")
                    print(f"      🔍 Key Strengths: {', '.join(diagnostic['key_strengths'])}")
                    if diagnostic['improvement_areas']:
                        print(f"      ⚠️  Areas for Improvement: {', '.join(diagnostic['improvement_areas'])}")
                    
                except Exception as e:
                    print(f"      ❌ Assessment failed: {e}")
                    import traceback
                    traceback.print_exc()
            
            detailed_results.append(job_results)
        
        # Generate comprehensive report
        self.generate_diagnostic_report(detailed_results)
        
        return detailed_results
    
    def generate_diagnostic_info(self, semantic_details: Dict, candidate: Dict, job: Dict) -> Dict:
        """Generate diagnostic information for the assessment"""
        
        # Determine match quality
        overall_score = semantic_details.get('final_calculation', {}).get('final_score_0_100', 0)
        
        if overall_score >= 80:
            match_quality = "Excellent Match"
        elif overall_score >= 70:
            match_quality = "Very Good Match"
        elif overall_score >= 60:
            match_quality = "Good Match"
        elif overall_score >= 50:
            match_quality = "Fair Match"
        else:
            match_quality = "Poor Match"
        
        # Identify key strengths
        key_strengths = []
        if semantic_details.get('education_relevance', 0) > 0.7:
            key_strengths.append("Strong Educational Background")
        if semantic_details.get('experience_relevance', 0) > 0.7:
            key_strengths.append("Relevant Work Experience")
        if semantic_details.get('skills_relevance', 0) > 0.7:
            key_strengths.append("Matching Skills")
        if semantic_details.get('overall_similarity', 0) > 0.8:
            key_strengths.append("Excellent Overall Profile Match")
        
        if not key_strengths:
            key_strengths.append("Profile shows potential")
        
        # Identify improvement areas
        improvement_areas = []
        if semantic_details.get('education_relevance', 0) < 0.3:
            improvement_areas.append("Educational Qualifications")
        if semantic_details.get('experience_relevance', 0) < 0.3:
            improvement_areas.append("Relevant Experience")
        if semantic_details.get('skills_relevance', 0) < 0.3:
            improvement_areas.append("Technical Skills")
        
        return {
            'match_quality': match_quality,
            'key_strengths': key_strengths,
            'improvement_areas': improvement_areas,
            'overall_score': overall_score,
            'content_quality_assessment': self.assess_content_quality(candidate)
        }
    
    def assess_content_quality(self, candidate: Dict) -> str:
        """Assess the quality of extracted candidate content"""
        education_length = len(candidate.get('education', ''))
        experience_length = len(candidate.get('experience', ''))
        skills_length = len(candidate.get('skills', ''))
        
        total_content = education_length + experience_length + skills_length
        
        if total_content > 500:
            return "Rich Content Profile"
        elif total_content > 200:
            return "Adequate Content Profile"
        else:
            return "Limited Content Profile"
    
    def generate_diagnostic_report(self, detailed_results: List[Dict]):
        """Generate comprehensive diagnostic report"""
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        report_filename = f"diagnostic_semantic_test_{timestamp}.json"
        
        # Calculate overall statistics
        all_scores = []
        assessment_insights = []
        
        for job_result in detailed_results:
            for candidate_assessment in job_result['candidate_assessments']:
                score = candidate_assessment['assessment_result'].get('semantic_score', 0)
                all_scores.append(score)
                
                assessment_insights.append({
                    'job_title': job_result['job_info']['title'],
                    'candidate_name': candidate_assessment['candidate_info']['name'],
                    'semantic_score': score,
                    'match_quality': candidate_assessment['diagnostic_info']['match_quality'],
                    'key_strengths': candidate_assessment['diagnostic_info']['key_strengths'],
                    'content_quality': candidate_assessment['diagnostic_info']['content_quality_assessment']
                })
        
        # Create comprehensive report
        diagnostic_report = {
            'test_metadata': {
                'test_date': datetime.now().isoformat(),
                'test_type': 'Diagnostic Real-World Semantic Assessment',
                'semantic_model': 'all-MiniLM-L6-v2',
                'total_assessments': len(all_scores)
            },
            'overall_statistics': {
                'average_score': sum(all_scores) / len(all_scores) if all_scores else 0,
                'highest_score': max(all_scores) if all_scores else 0,
                'lowest_score': min(all_scores) if all_scores else 0,
                'score_distribution': {
                    'excellent_80_plus': len([s for s in all_scores if s >= 80]),
                    'very_good_70_79': len([s for s in all_scores if 70 <= s < 80]),
                    'good_60_69': len([s for s in all_scores if 60 <= s < 70]),
                    'fair_50_59': len([s for s in all_scores if 50 <= s < 60]),
                    'poor_below_50': len([s for s in all_scores if s < 50])
                }
            },
            'assessment_insights': assessment_insights,
            'detailed_results': detailed_results,
            'recommendations': self.generate_recommendations(assessment_insights)
        }
        
        # Save report
        with open(report_filename, 'w', encoding='utf-8') as f:
            json.dump(diagnostic_report, f, indent=2, ensure_ascii=False)
        
        # Print summary
        print(f"\n📊 DIAGNOSTIC TEST REPORT")
        print("=" * 50)
        print(f"📅 Test Completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"🧪 Total Assessments: {len(all_scores)}")
        
        if all_scores:
            print(f"\n📈 Score Analysis:")
            print(f"   Average Score: {diagnostic_report['overall_statistics']['average_score']:.1f}")
            print(f"   Score Range: {diagnostic_report['overall_statistics']['lowest_score']:.1f} - {diagnostic_report['overall_statistics']['highest_score']:.1f}")
            
            dist = diagnostic_report['overall_statistics']['score_distribution']
            print(f"\n📊 Performance Distribution:")
            print(f"   Excellent (80+): {dist['excellent_80_plus']} assessments")
            print(f"   Very Good (70-79): {dist['very_good_70_79']} assessments")
            print(f"   Good (60-69): {dist['good_60_69']} assessments")
            print(f"   Fair (50-59): {dist['fair_50_59']} assessments")
            print(f"   Needs Improvement (<50): {dist['poor_below_50']} assessments")
        
        print(f"\n🏆 Best Matches:")
        sorted_insights = sorted(assessment_insights, key=lambda x: x['semantic_score'], reverse=True)
        for insight in sorted_insights[:3]:
            print(f"   {insight['candidate_name']} → {insight['job_title']}: {insight['semantic_score']:.1f} ({insight['match_quality']})")
        
        print(f"\n📁 Detailed report saved: {report_filename}")
        
        return diagnostic_report
    
    def generate_recommendations(self, assessment_insights: List[Dict]) -> List[str]:
        """Generate recommendations based on assessment results"""
        recommendations = []
        
        # Score-based recommendations
        scores = [insight['semantic_score'] for insight in assessment_insights]
        avg_score = sum(scores) / len(scores) if scores else 0
        
        if avg_score < 30:
            recommendations.append("Consider improving PDS parsing algorithms to extract more relevant content")
            recommendations.append("Review semantic scoring weights to better match university job requirements")
        elif avg_score < 50:
            recommendations.append("PDS content extraction is working but may need fine-tuning for better relevance")
        else:
            recommendations.append("Semantic scoring system is performing well with current PDS files")
        
        # Content quality recommendations
        content_qualities = [insight['content_quality'] for insight in assessment_insights]
        limited_content = content_qualities.count('Limited Content Profile')
        
        if limited_content > len(content_qualities) * 0.5:
            recommendations.append("Many PDS files have limited extractable content - consider improving parsing")
        
        return recommendations

def main():
    """Run the diagnostic test"""
    print("🚀 Starting Diagnostic Real-World Semantic Test")
    print("=" * 60)
    
    try:
        tester = DiagnosticSemanticTest()
        results = tester.run_diagnostic_test()
        
        print("\n🎉 Diagnostic test completed successfully!")
        print("✅ Semantic scoring system thoroughly analyzed with real PDS files")
        
        return True
        
    except Exception as e:
        print(f"\n❌ Test failed: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)