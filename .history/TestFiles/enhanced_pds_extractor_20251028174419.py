#!/usr/bin/env python3
"""
Enhanced PDS Content Extractor
Better parsing of PDS files to extract meaningful candidate information
"""

import os
import sys
import json
import openpyxl
from typing import Dict, Optional, List
import re

def extract_comprehensive_pds_content(filepath: str, filename: str) -> Optional[Dict]:
    """
    Enhanced PDS content extraction with better structure analysis
    """
    try:
        workbook = openpyxl.load_workbook(filepath, data_only=True)
        
        # Look through all sheets
        candidate_info = {
            'filename': filename,
            'name': 'Unknown Candidate',
            'education': '',
            'experience': '',
            'skills': '',
            'extracted_text': '',
            'debug_info': []
        }
        
        # Process each sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            candidate_info['debug_info'].append(f"Processing sheet: {sheet_name}")
            
            # Extract all meaningful content
            all_content = []
            potential_names = []
            education_content = []
            experience_content = []
            skills_content = []
            
            # Scan the entire sheet
            max_row = min(sheet.max_row, 200)  # Limit for performance
            max_col = min(sheet.max_column, 20)
            
            for row in range(1, max_row + 1):
                row_content = []
                
                for col in range(1, max_col + 1):
                    try:
                        cell_value = sheet.cell(row=row, column=col).value
                        if cell_value and isinstance(cell_value, str):
                            text = cell_value.strip()
                            if len(text) > 2:
                                row_content.append(text)
                                all_content.append(text)
                                
                                # Look for names (specific patterns)
                                if is_potential_name(text):
                                    potential_names.append((text, row, col))
                                
                                # Look for education keywords
                                if contains_education_keywords(text):
                                    education_content.append(text)
                                
                                # Look for experience keywords
                                if contains_experience_keywords(text):
                                    experience_content.append(text)
                                
                                # Look for skills/training keywords
                                if contains_skills_keywords(text):
                                    skills_content.append(text)
                    except:
                        continue
                
                # Log some row content for debugging
                if len(row_content) > 0 and row <= 50:
                    candidate_info['debug_info'].append(f"Row {row}: {row_content[:3]}")
        
        workbook.close()
        
        # Find best name
        best_name = find_best_candidate_name(potential_names, all_content)
        if best_name:
            candidate_info['name'] = best_name
        
        # Compile sections
        candidate_info['education'] = compile_section_content(education_content)
        candidate_info['experience'] = compile_section_content(experience_content)
        candidate_info['skills'] = compile_section_content(skills_content)
        candidate_info['extracted_text'] = ' '.join(all_content[:100])  # First 100 items
        
        return candidate_info
        
    except Exception as e:
        print(f"Error processing {filename}: {e}")
        return None

def is_potential_name(text: str) -> bool:
    """Check if text could be a person's name"""
    if len(text) < 5 or len(text) > 60:
        return False
    
    # Must contain spaces (first + last name)
    if ' ' not in text:
        return False
    
    # Should be mostly alphabetic
    alpha_ratio = sum(c.isalpha() or c.isspace() for c in text) / len(text)
    if alpha_ratio < 0.8:
        return False
    
    # Should have 2-4 words
    words = text.split()
    if len(words) < 2 or len(words) > 4:
        return False
    
    # Each word should be reasonable length
    if any(len(word) < 2 or len(word) > 15 for word in words):
        return False
    
    # Avoid common form field text
    exclude_patterns = [
        'personal data sheet', 'civil service', 'government', 'position',
        'date of birth', 'place of birth', 'sex', 'citizenship', 'height',
        'weight', 'blood type', 'gsis', 'pag-ibig', 'philhealth', 'sss',
        'tin', 'agency', 'employee', 'yes', 'no', 'male', 'female',
        'single', 'married', 'widowed', 'separated', 'elementary',
        'secondary', 'vocational', 'college', 'graduate', 'bachelor',
        'master', 'doctor', 'course', 'degree', 'school', 'university'
    ]
    
    text_lower = text.lower()
    if any(pattern in text_lower for pattern in exclude_patterns):
        return False
    
    # Prefer title case
    if text.istitle():
        return True
    
    # Check if it looks like a name pattern
    name_pattern = re.match(r'^[A-Z][a-z]+ (?:[A-Z][a-z]*\.? )?[A-Z][a-z]+$', text)
    return bool(name_pattern)

def contains_education_keywords(text: str) -> bool:
    """Check if text contains education-related content"""
    education_keywords = [
        'education', 'school', 'university', 'college', 'degree', 'bachelor',
        'master', 'phd', 'doctorate', 'graduate', 'undergraduate', 'course',
        'major', 'minor', 'gpa', 'cum laude', 'magna cum laude', 'summa cum laude',
        'elementary', 'secondary', 'high school', 'vocational', 'technical',
        'diploma', 'certificate', 'academic', 'honors', 'scholarship'
    ]
    
    text_lower = text.lower()
    return any(keyword in text_lower for keyword in education_keywords)

def contains_experience_keywords(text: str) -> bool:
    """Check if text contains work experience content"""
    experience_keywords = [
        'experience', 'work', 'employment', 'job', 'position', 'title',
        'company', 'organization', 'employer', 'supervisor', 'manager',
        'director', 'coordinator', 'assistant', 'associate', 'specialist',
        'analyst', 'officer', 'administrator', 'consultant', 'developer',
        'engineer', 'teacher', 'instructor', 'professor', 'researcher',
        'years', 'months', 'full-time', 'part-time', 'contract', 'internship'
    ]
    
    text_lower = text.lower()
    return any(keyword in text_lower for keyword in experience_keywords)

def contains_skills_keywords(text: str) -> bool:
    """Check if text contains skills/training content"""
    skills_keywords = [
        'skills', 'training', 'certification', 'seminar', 'workshop',
        'conference', 'course', 'program', 'development', 'competency',
        'proficiency', 'expertise', 'knowledge', 'ability', 'capability',
        'programming', 'software', 'computer', 'technology', 'technical',
        'language', 'communication', 'leadership', 'management', 'research',
        'analysis', 'microsoft', 'excel', 'word', 'powerpoint', 'office'
    ]
    
    text_lower = text.lower()
    return any(keyword in text_lower for keyword in skills_keywords)

def find_best_candidate_name(potential_names: List[tuple], all_content: List[str]) -> str:
    """Find the most likely candidate name from potential matches"""
    if not potential_names:
        return "Unknown Candidate"
    
    # Score each potential name
    scored_names = []
    
    for name, row, col in potential_names:
        score = 0
        
        # Prefer names that appear early in the document
        if row <= 20:
            score += 5
        elif row <= 50:
            score += 3
        
        # Prefer names in the first few columns (likely to be labels)
        if col <= 5:
            score += 3
        
        # Prefer proper title case
        if name.istitle():
            score += 4
        
        # Prefer reasonable length
        if 10 <= len(name) <= 30:
            score += 2
        
        # Prefer 2-3 words (first + last, or first + middle + last)
        word_count = len(name.split())
        if word_count == 2:
            score += 3
        elif word_count == 3:
            score += 4
        
        # Check if it appears multiple times (good sign)
        appearances = sum(1 for content in all_content if name in content)
        if appearances > 1:
            score += appearances
        
        scored_names.append((score, name))
    
    # Sort by score and return best
    scored_names.sort(reverse=True)
    return scored_names[0][1] if scored_names else "Unknown Candidate"

def compile_section_content(content_list: List[str]) -> str:
    """Compile section content into a meaningful summary"""
    if not content_list:
        return "Information not clearly extractable from PDS format"
    
    # Remove duplicates while preserving order
    unique_content = []
    seen = set()
    
    for item in content_list:
        if item not in seen and len(item) > 10:  # Only meaningful content
            unique_content.append(item)
            seen.add(item)
    
    # Combine into a readable summary
    if unique_content:
        return '. '.join(unique_content[:5])  # Top 5 items
    else:
        return "Relevant information found but requires manual review"

def test_pds_extraction():
    """Test the enhanced PDS extraction"""
    pds_folder = r"C:\Users\Lenar Yolola\OneDrive\Desktop\ResuAI__\SamplePDSFiles"
    
    if not os.path.exists(pds_folder):
        print(f"❌ PDS folder not found: {pds_folder}")
        return
    
    pds_files = [f for f in os.listdir(pds_folder) if f.endswith('.xlsx')]
    print(f"🔍 Testing enhanced extraction on {len(pds_files)} files")
    print("=" * 60)
    
    for filename in pds_files:
        filepath = os.path.join(pds_folder, filename)
        print(f"\n📄 Processing: {filename}")
        
        result = extract_comprehensive_pds_content(filepath, filename)
        
        if result:
            print(f"   👤 Name: {result['name']}")
            print(f"   📚 Education: {result['education'][:100]}...")
            print(f"   💼 Experience: {result['experience'][:100]}...")
            print(f"   🛠️  Skills: {result['skills'][:100]}...")
            print(f"   📝 Content Length: {len(result['extracted_text'])} chars")
            
            # Show some debug info
            print(f"   🔍 Debug Info (first 5):")
            for info in result['debug_info'][:5]:
                print(f"      {info}")
        else:
            print(f"   ❌ Failed to extract content")

if __name__ == "__main__":
    test_pds_extraction()