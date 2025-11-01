#!/usr/bin/env python3
"""
Enhanced PDS Extractor with Phase 2 fixes
Addresses:
1. Improved personal info extraction
2. Civil service eligibility validation
3. Reference validation
4. Section contamination prevention
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import re
import os
from datetime import datetime
import json
from typing import List, Dict, Any, Optional
from improved_pds_extractor import ImprovedPDSExtractor

class EnhancedPDSExtractor(ImprovedPDSExtractor):
    """Enhanced extractor with Phase 2 fixes"""
    
    def __init__(self):
        super().__init__()
        # Enhanced patterns for better extraction
        self.name_patterns = [
            r'SURNAME[:\s]*([A-Za-z\s]+)',
            r'FIRST NAME[:\s]*([A-Za-z\s]+)',
            r'MIDDLE NAME[:\s]*([A-Za-z\s]+)',
            # Alternative patterns
            r'Last Name[:\s]*([A-Za-z\s]+)',
            r'Given Name[:\s]*([A-Za-z\s]+)',
        ]
        
        self.phone_patterns = [
            r'MOBILE NO[:\.\s]*([0-9\+\-\(\)\s]{10,15})',
            r'TELEPHONE NO[:\.\s]*([0-9\+\-\(\)\s]{7,15})',
            r'PHONE[:\.\s]*([0-9\+\-\(\)\s]{7,15})',
            r'CONTACT[:\.\s]*([0-9\+\-\(\)\s]{7,15})',
            # Pattern for Philippine mobile numbers
            r'(09[0-9]{9})',
            r'(\+639[0-9]{9})',
        ]
        
        self.eligibility_keywords = [
            'civil service', 'eligibility', 'examination', 'board exam', 
            'professional regulation commission', 'prc', 'licensure',
            'career service', 'rating', 'board rating'
        ]
        
        self.work_experience_keywords = [
            'company', 'corporation', 'inc', 'office', 'department',
            'position', 'salary', 'supervisor', 'employment', 'work'
        ]
    
    def _extract_personal_info_enhanced(self, worksheet) -> Dict[str, Any]:
        """Enhanced personal info extraction with better patterns"""
        personal_info = {}
        
        try:
            # Use the original method first
            personal_info = super()._extract_personal_info(worksheet)
            
            # Enhanced name extraction if original failed
            if not personal_info.get('full_name') or personal_info.get('full_name') == '':
                name_data = self._extract_name_enhanced(worksheet)
                personal_info.update(name_data)
            
            # Enhanced phone extraction
            phone_data = self._extract_phone_enhanced(worksheet)
            personal_info.update(phone_data)
            
            # Clean and validate
            personal_info = self._validate_personal_info(personal_info)
            
        except Exception as e:
            self.errors.append(f"Enhanced personal info extraction error: {str(e)}")
        
        return personal_info
    
    def _extract_name_enhanced(self, worksheet) -> Dict[str, Any]:
        """Enhanced name extraction with multiple patterns"""
        name_data = {}
        
        try:
            # Search entire worksheet for name patterns
            for row in range(1, min(30, worksheet.max_row + 1)):  # Search first 30 rows
                for col in range(1, min(10, worksheet.max_column + 1)):  # Search first 10 columns
                    cell_value = worksheet.cell(row=row, column=col).value
                    if not cell_value:
                        continue
                    
                    cell_text = str(cell_value).strip().upper()
                    
                    # Check for surname
                    if 'SURNAME' in cell_text and not name_data.get('surname'):
                        # Look in adjacent cells
                        for offset in [1, 2, 3]:
                            if col + offset <= worksheet.max_column:
                                adjacent_value = worksheet.cell(row=row, column=col + offset).value
                                if adjacent_value and str(adjacent_value).strip():
                                    candidate_surname = str(adjacent_value).strip()
                                    if self._is_valid_name_part(candidate_surname):
                                        name_data['surname'] = candidate_surname
                                        break
                    
                    # Check for first name
                    if 'FIRST NAME' in cell_text and not name_data.get('first_name'):
                        for offset in [1, 2, 3]:
                            if col + offset <= worksheet.max_column:
                                adjacent_value = worksheet.cell(row=row, column=col + offset).value
                                if adjacent_value and str(adjacent_value).strip():
                                    candidate_first = str(adjacent_value).strip()
                                    if self._is_valid_name_part(candidate_first):
                                        name_data['first_name'] = candidate_first
                                        break
                    
                    # Check for middle name
                    if 'MIDDLE NAME' in cell_text and not name_data.get('middle_name'):
                        for offset in [1, 2, 3]:
                            if col + offset <= worksheet.max_column:
                                adjacent_value = worksheet.cell(row=row, column=col + offset).value
                                if adjacent_value and str(adjacent_value).strip():
                                    candidate_middle = str(adjacent_value).strip()
                                    if self._is_valid_name_part(candidate_middle):
                                        name_data['middle_name'] = candidate_middle
                                        break
            
            # Create full name
            if name_data.get('first_name') or name_data.get('surname'):
                name_parts = []
                if name_data.get('first_name'):
                    name_parts.append(name_data['first_name'])
                if name_data.get('middle_name'):
                    name_parts.append(name_data['middle_name'])
                if name_data.get('surname'):
                    name_parts.append(name_data['surname'])
                
                name_data['full_name'] = ' '.join(name_parts)
                name_data['name'] = name_data['full_name']  # Also set 'name' field
                
        except Exception as e:
            self.warnings.append(f"Enhanced name extraction error: {str(e)}")
        
        return name_data
    
    def _extract_phone_enhanced(self, worksheet) -> Dict[str, Any]:
        """Enhanced phone number extraction"""
        phone_data = {}
        
        try:
            # Search for phone patterns in worksheet
            for row in range(1, min(30, worksheet.max_row + 1)):
                for col in range(1, min(15, worksheet.max_column + 1)):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if not cell_value:
                        continue
                    
                    cell_text = str(cell_value).strip()
                    
                    # Check for mobile/phone numbers
                    phone_match = re.search(r'(09[0-9]{9}|\+639[0-9]{9}|[0-9]{7,11})', cell_text)
                    if phone_match:
                        phone_number = phone_match.group(1)
                        
                        # Determine if it's mobile or telephone
                        if phone_number.startswith('09') or phone_number.startswith('+639'):
                            if not phone_data.get('mobile_no'):
                                phone_data['mobile_no'] = phone_number
                                phone_data['phone'] = phone_number  # Also set generic 'phone' field
                        elif len(phone_number) >= 7:
                            if not phone_data.get('telephone_no'):
                                phone_data['telephone_no'] = phone_number
                                if not phone_data.get('phone'):  # Use as generic phone if mobile not found
                                    phone_data['phone'] = phone_number
                        
        except Exception as e:
            self.warnings.append(f"Enhanced phone extraction error: {str(e)}")
        
        return phone_data
    
    def _is_valid_name_part(self, text: str) -> bool:
        """Validate if text looks like a name part"""
        if not text or len(text.strip()) < 2:
            return False
        
        # Check if it contains only letters, spaces, periods, and common name characters
        if not re.match(r'^[A-Za-z\s\.\'-]+$', text.strip()):
            return False
        
        # Exclude common non-name words
        exclude_words = ['NONE', 'N/A', 'NOT APPLICABLE', 'PERSONAL', 'DATA', 'SHEET']
        if text.strip().upper() in exclude_words:
            return False
        
        return True
    
    def _validate_personal_info(self, personal_info: Dict[str, Any]) -> Dict[str, Any]:
        """Validate and clean personal information"""
        cleaned = {}
        
        for key, value in personal_info.items():
            if value and str(value).strip() and str(value).strip() != 'None':
                cleaned_value = str(value).strip()
                
                # Additional validation for specific fields
                if key in ['email']:
                    if '@' in cleaned_value and '.' in cleaned_value:
                        cleaned[key] = cleaned_value
                elif key in ['mobile_no', 'telephone_no', 'phone']:
                    if re.match(r'^[0-9\+\-\(\)\s]{7,15}$', cleaned_value):
                        cleaned[key] = cleaned_value
                else:
                    cleaned[key] = cleaned_value
        
        return cleaned
    
    def _extract_civil_service_eligibility_enhanced(self, worksheet) -> List[Dict[str, Any]]:
        """Enhanced civil service eligibility with contamination prevention"""
        eligibility_entries = []
        
        try:
            # Use original method first
            original_entries = super()._extract_civil_service_eligibility(worksheet)
            
            # Filter and validate each entry
            for entry in original_entries:
                if self._is_valid_eligibility_entry(entry):
                    eligibility_entries.append(entry)
                else:
                    self.warnings.append(f"Filtered invalid eligibility entry: {str(entry)[:100]}")
        
        except Exception as e:
            self.errors.append(f"Enhanced eligibility extraction error: {str(e)}")
        
        return eligibility_entries
    
    def _is_valid_eligibility_entry(self, entry: Dict[str, Any]) -> bool:
        """Validate if entry is actually a civil service eligibility"""
        if not isinstance(entry, dict):
            return False
        
        eligibility_text = entry.get('eligibility', '').lower()
        
        # Must have eligibility text
        if not eligibility_text or len(eligibility_text.strip()) < 3:
            return False
        
        # Check if it contains eligibility-related keywords
        has_eligibility_keywords = any(keyword in eligibility_text for keyword in self.eligibility_keywords)
        
        # Check if it contains work experience keywords (bad sign)
        has_work_keywords = any(keyword in eligibility_text for keyword in self.work_experience_keywords)
        
        # Reject if it has work keywords but no eligibility keywords
        if has_work_keywords and not has_eligibility_keywords:
            return False
        
        # Additional validation: check other fields
        entry_text = json.dumps(entry).lower()
        work_contamination_score = sum(1 for keyword in self.work_experience_keywords if keyword in entry_text)
        
        # If too many work-related terms, probably contaminated
        if work_contamination_score > 2:
            return False
        
        return True
    
    def _extract_other_information_enhanced(self, worksheet) -> Dict[str, Any]:
        """Enhanced other information extraction with reference validation"""
        other_info = super()._extract_other_information(worksheet)
        
        # Validate and clean references
        if 'references' in other_info:
            valid_references = []
            for ref in other_info['references']:
                if self._is_valid_reference(ref):
                    valid_references.append(ref)
                else:
                    self.warnings.append(f"Filtered invalid reference: {str(ref)[:100]}")
            
            other_info['references'] = valid_references
        
        return other_info
    
    def _is_valid_reference(self, reference: Dict[str, Any]) -> bool:
        """Validate if reference entry is actually a person"""
        if not isinstance(reference, dict):
            return False
        
        name = reference.get('name', '').strip()
        
        # Must have a name
        if not name or len(name) < 3:
            return False
        
        # Check if it looks like a person's name
        # Should contain letters and spaces, possibly periods
        if not re.match(r'^[A-Za-z\s\.\'-]+$', name):
            return False
        
        # Should have at least 2 parts (first and last name)
        name_parts = name.split()
        if len(name_parts) < 2:
            return False
        
        # Exclude obvious non-names
        exclude_patterns = [
            r'\d+',  # Contains numbers
            r'(office|department|company|corporation|inc)',  # Business terms
            r'(address|telephone|email|contact)',  # Contact field labels
        ]
        
        for pattern in exclude_patterns:
            if re.search(pattern, name.lower()):
                return False
        
        return True
    
    # Override main extraction method to use enhanced versions
    def _extract_from_excel(self, file_path: str) -> Dict[str, Any]:
        """Enhanced Excel extraction using improved methods"""
        try:
            workbook = load_workbook(file_path, data_only=True)
            
            # Extract from different sheets
            if 'C1' in workbook.sheetnames:
                c1_sheet = workbook['C1']
                self.pds_data['personal_info'] = self._extract_personal_info_enhanced(c1_sheet)
                self.pds_data['family_background'] = self._extract_family_background(c1_sheet)
                self.pds_data['educational_background'] = self._extract_educational_background(c1_sheet)
            
            if 'C2' in workbook.sheetnames:
                c2_sheet = workbook['C2']
                self.pds_data['civil_service_eligibility'] = self._extract_civil_service_eligibility_enhanced(c2_sheet)
                self.pds_data['work_experience'] = self._extract_work_experience(c2_sheet)
            
            if 'C3' in workbook.sheetnames:
                c3_sheet = workbook['C3']
                self.pds_data['voluntary_work'] = self._extract_voluntary_work(c3_sheet)
                self.pds_data['learning_development'] = self._extract_learning_development(c3_sheet)
                
            if 'C4' in workbook.sheetnames:
                c4_sheet = workbook['C4']
                self.pds_data['other_information'] = self._extract_other_information_enhanced(c4_sheet)
            
            # If no C1-C4 sheets, try first sheet
            if not any(sheet in workbook.sheetnames for sheet in ['C1', 'C2', 'C3', 'C4']):
                first_sheet = workbook.active
                self.pds_data['personal_info'] = self._extract_personal_info_enhanced(first_sheet)
                self.pds_data['family_background'] = self._extract_family_background(first_sheet)
                self.pds_data['educational_background'] = self._extract_educational_background(first_sheet)
                self.pds_data['civil_service_eligibility'] = self._extract_civil_service_eligibility_enhanced(first_sheet)
                self.pds_data['work_experience'] = self._extract_work_experience(first_sheet)
                self.pds_data['voluntary_work'] = self._extract_voluntary_work(first_sheet)
                self.pds_data['learning_development'] = self._extract_learning_development(first_sheet)
                self.pds_data['other_information'] = self._extract_other_information_enhanced(first_sheet)
            
            # Add extraction metadata
            self.pds_data['extraction_metadata'] = {
                'file_type': 'XLSX',
                'extraction_method': 'enhanced_table_parsing',
                'extracted_at': datetime.now().isoformat(),
                'sheets_found': workbook.sheetnames,
                'errors': self.errors,
                'warnings': self.warnings,
                'enhancement_version': 'phase2_v1.0'
            }
            
            workbook.close()
            return self.pds_data
            
        except Exception as e:
            self.errors.append(f"Enhanced Excel extraction error: {str(e)}")
            return {}
